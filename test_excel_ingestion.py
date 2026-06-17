import openpyxl

def detect_headers(sheet, max_scan=10):
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
        if len(non_empty) >= 2:  # likely header row
            headers = [str(cell).strip() if cell else f"Column{j}" for j, cell in enumerate(row)]
            return i, headers
    return 1, []

def main():
    file_path = r"C:\Users\Immanuel\Desktop\Project-Management-Sample-Data.xlsx"
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    header_row_index, headers = detect_headers(sheet)

    # ✅ Drop Column0 if it’s just a placeholder
    if headers and headers[0].startswith("Column"):
        headers = headers[1:]

    rows = []
    for row in sheet.iter_rows(min_row=header_row_index+1, values_only=True):
        if not any(row):
            continue
        # ✅ Drop the first cell to align with headers
        row = row[1:]
        row_dict = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            row_dict[header] = value
        rows.append(row_dict)

    print("Headers:", headers)
    print("Total rows:", len(rows))
    print("First 3 rows:")
    for r in rows[:3]:
        print(r)

if __name__ == "__main__":
    main()
