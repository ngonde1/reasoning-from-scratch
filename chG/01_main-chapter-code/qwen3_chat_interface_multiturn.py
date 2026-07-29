# --- Standard library imports ---
import os
import asyncio
import time
import datetime
import csv
import re
import json
from pathlib import Path
from typing import TypedDict, List
from pdf2image import convert_from_path
import langcodes
from fastapi.staticfiles import StaticFiles
from local_storage_client import LocalStorageClient
from sqlalchemy import insert
from chainlit.context import context
from sqlalchemy import insert
from chainlit.context import context
from collections import defaultdict
from sqlalchemy import text as sql_text
from llama_parse import LlamaParse
from docling.document_converter import DocumentConverter
from reasoning_from_scratch.rag_utils import add_to_index, query_index
#from chainlit.element import DataTable
from reasoning_from_scratch.rag_utils import add_to_index
from reasoning_from_scratch.utils import sanitize_row, chunk_rows, chunk_text, ingest_file, get_data_layer


# ✅ Initialize advanced parsers
parser = LlamaParse(api_key=os.getenv("LLAMA_PARSE_API_KEY"))
doc_converter = DocumentConverter()


# ✅ Import your universal ingestion helpers
#from file_ingestion import (
 #   load_excel, load_csv, load_docx, load_pdf,
  #  load_txt, load_json,
   # auto_detect_columns, build_copy_table, compute_all_numeric_stats,
    #professional_llm_response
#)


# Toggle streaming mode
USE_STREAMING = True   # for live typing
# USE_STREAMING = False  # for full polished answer

# --- Third-party libraries ---
import torch
import chainlit as cl
from chainlit.types import ThreadDict
from chainlit.data.sql_alchemy import SQLAlchemyDataLayer
from langgraph.graph import StateGraph
from langchain_core.messages import HumanMessage, BaseMessage, AIMessage
from mcp import ClientSession  # ✅ MCP import
import socketio  # ✅ Added for payload limit
import faiss
import numpy as np
from sentence_transformers import SentenceTransformer
from sentence_transformers import SentenceTransformer
from dotenv import load_dotenv
load_dotenv()

# Now os.getenv will work
parser = LlamaParse(api_key=os.getenv("LLAMA_PARSE_API_KEY"))


# --- Chainlit elements ---
from chainlit import Image, Pdf

# --- File handling libraries ---
import PyPDF2
import docx
import openpyxl
from PIL import Image as PILImage
import pytesseract
from langchain_community.document_loaders import PDFPlumberLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter

def sanitize_row(row: dict) -> dict:
    """Convert non-serializable values (like datetime) into strings."""
    clean_row = {}
    for k, v in row.items():
        if isinstance(v, (datetime.date, datetime.datetime)):
            clean_row[k] = v.isoformat()  # e.g., '2021-10-16T00:00:00'
        elif v is None:
            clean_row[k] = ""
        else:
            clean_row[k] = v
    return clean_row


# --- Local project imports ---
from reasoning_from_scratch.ch02 import get_device, generate_text_basic_stream_cache
from reasoning_from_scratch.ch03 import load_model_and_tokenizer, load_tokenizer_only
from reasoning_from_scratch.qwen3 import Qwen3Model, QWEN_CONFIG_06_B

# ------------------- Socket.IO Payload Limit -------------------
# Increase payload limit to 200 MB
sio = socketio.AsyncServer(
    async_mode="asgi",
    max_http_buffer_size=200 * 1024 * 1024  # 200 MB
)

# ------------------- Configuration -------------------
WHICH_MODEL = "reasoning"
MAX_NEW_TOKENS = 38912
LOCAL_DIR = "qwen3"
CHECKPOINT_PATH = os.getenv("CHECKPOINT_PATH")
COMPILE = False

DEVICE = get_device()

def load_app_model_and_tokenizer():
    if CHECKPOINT_PATH is None:
        return load_model_and_tokenizer(
            which_model=WHICH_MODEL,
            device=DEVICE,
            use_compile=COMPILE,
            local_dir=LOCAL_DIR,
        )
    checkpoint_path = Path(CHECKPOINT_PATH)
    if not checkpoint_path.exists():
        raise FileNotFoundError(f"Checkpoint file not found: {checkpoint_path}")
    tokenizer = load_tokenizer_only(which_model=WHICH_MODEL, local_dir=LOCAL_DIR)
    model = Qwen3Model(QWEN_CONFIG_06_B)
    model.load_state_dict(torch.load(checkpoint_path, map_location="cpu"))
    model.to(DEVICE)
    if COMPILE:
        torch._dynamo.config.allow_unspec_int_on_nn_module = True
        model = torch.compile(model)
    return model, tokenizer

MODEL, TOKENIZER, DEVICE = load_app_model_and_tokenizer()

EOS_TOKEN_IDS = (
    TOKENIZER.encode("<|im_end|>")[0],
    TOKENIZER.encode("<|endoftext|>")[0]
)

# ============================================================
# Conversation History Tracker
# ============================================================
def build_prompt_from_history(history, add_assistant_header=True):
    parts = []
    for m in history:
        role = m["role"]
        content = m["content"]
        parts.append(f"<|im_start|>{role}\n{content}<|im_end|>\n")
    if add_assistant_header:
        parts.append("<|im_start|>assistant\n")
    return "".join(parts)

# ------------------- Agent State -------------------
class AgentState(TypedDict):
    messages: List[BaseMessage]

# ------------------- Sync Qwen3 runner -------------------
def run_qwen_sync(prompt: str) -> str:
    input_ids = TOKENIZER.encode(prompt)
    max_context = MODEL.cfg["context_length"]

    # Truncate safely to avoid exceeding context length
    if len(input_ids) > max_context:
        input_ids = input_ids[-max_context:]

    input_ids_tensor = torch.tensor(input_ids, device=DEVICE).unsqueeze(0)

    output = ""
    for tok in generate_text_basic_stream_cache(
        model=MODEL,
        token_ids=input_ids_tensor,   # <-- pass raw IDs, not embeddings
        max_new_tokens=MAX_NEW_TOKENS,
    ):
        token_id = tok.squeeze(0)
        if token_id in EOS_TOKEN_IDS:
            break
        output += TOKENIZER.decode(token_id.tolist())
    return output

# ------------------- Authentication & Data Layer -------------------
@cl.password_auth_callback
def auth_callback(username: str, password: str):
    if (username, password) == ("admin", "kevin"):
        return cl.User(identifier="admin", metadata={"role": "admin", "provider": "credentials"})
    return None


# Use Chainlit’s FastAPI app instead of cl.app
app = cl.server.app
app.mount("/files", StaticFiles(directory="uploads"), name="files")

@cl.data_layer
def get_data_layer():
    conninfo = os.getenv("DATABASE_URL")
    if not conninfo:
        raise ValueError("DATABASE_URL not found in environment variables.")
    return SQLAlchemyDataLayer(
        conninfo=conninfo,
        storage_provider=LocalStorageClient(base_dir="uploads")
    )

# ------------------- Stripe Toggle -------------------
def register_tools(app):
    if os.getenv("STRIPE_ENABLED", "false").lower() == "true":
        try:
            from stripe_mcp import StripeMCP
            app.register_tool(StripeMCP())
            print("✅ Stripe MCP enabled.")
        except ImportError:
            print("⚠️ Stripe MCP module not installed — skipping Stripe integration.")
    else:
        print("🚫 Stripe MCP disabled — prompt-only mode.")

# ------------------- MCP Integration -------------------
@cl.on_mcp_connect
async def on_mcp_connect(connection, session: ClientSession):
    result = await session.list_tools()
    tools = [{
        "name": t.name,
        "description": t.description,
        "input_schema": t.inputSchema,
    } for t in result.tools]
    cl.user_session.set("mcp_tools", {connection.name: tools})
    print(f"✅ Connected to {connection.name} with tools: {[t['name'] for t in tools]}")

@cl.on_mcp_disconnect
async def on_mcp_disconnect(name: str, session: ClientSession):
    print(f"❌ MCP connection {name} closed.")

@cl.step(type="tool")
async def call_tool(tool_use):
    tool_name = tool_use.name
    tool_input = tool_use.input
    mcp_tools = cl.user_session.get("mcp_tools", {})
    mcp_name = next((name for name, tools in mcp_tools.items() if any(t["name"] == tool_name for t in tools)), None)
    if not mcp_name:
        return {"error": f"Tool {tool_name} not found"}
    mcp_session, _ = cl.context.session.mcp_sessions.get(mcp_name)
    result = await mcp_session.call_tool(tool_name, tool_input)
    return result

# ------------------- Chat Lifecycle -------------------
@cl.on_chat_resume
async def on_chat_resume(thread: ThreadDict):
    try:
        # ✅ Restore conversation messages
        messages = []
        for m in thread.get("messages", []):
            role = m.get("role")
            content = m.get("content", "").strip()
            if not content:
                continue
            if role == "user":
                messages.append(HumanMessage(content=content))
            elif role == "assistant":
                messages.append(AIMessage(content=content))

        cl.user_session.set("state", {"messages": messages})
        cl.user_session.set("history", [
            {"role": m.__class__.__name__.replace("Message", "").lower(), "content": m.content}
            for m in messages
        ])

        # ✅ Reload file context from Postgres (no re‑persist)
        data_layer = get_data_layer()
        async with data_layer.engine.begin() as conn:
            result = await conn.execute(
                sql_text("SELECT file_name, headers, rows, text_content FROM file_contents WHERE thread_id = :tid"),
                {"tid": thread["id"]}
            )
            for row in result:
                if row.file_name:
                    headers, rows, text_content = [], [], ""

                    if row.headers:
                        headers = json.loads(row.headers) if isinstance(row.headers, str) else row.headers
                    if row.rows:
                        rows = json.loads(row.rows) if isinstance(row.rows, str) else row.rows
                    if row.text_content:
                        text_content = row.text_content

                    # ✅ Store back into user_session for later use
                    if headers:
                        cl.user_session.set(f"{row.file_name}_headers", headers)
                    if rows:
                        cl.user_session.set(f"{row.file_name}_table", rows)
                    if text_content:
                        cl.user_session.set(f"{row.file_name}_text", text_content)

                    print(f"✅ Reloaded {row.file_name} from Postgres")

    except Exception as e:
        print(f"⚠️ Error resuming chat: {e}")
        cl.user_session.set("state", {"messages": []})
        cl.user_session.set("history", [])


#@cl.on_file_upload
#async def on_file_upload(file):
 #   """
  #  Automatically ingest files uploaded mid-conversation.
   # """
 #   try:
  #      await persist_file_content(file.path)
   #     print(f"✅ Auto-ingested {file.path} during conversation")
    #    await cl.Message(
     #       content=f"✅ Your file **{Path(file.path).name}** has been ingested and indexed."
      #  ).send()
 #   except Exception as e:
  #      print(f"⚠️ Failed to ingest {file.path}: {e}")
   #     await cl.Message(
    #        content=f"⚠️ Failed to ingest {Path(file.path).name}: {e}"
     #   ).send()

@cl.on_chat_start
async def on_chat_start():
    cl.user_session.set("history", [])
    cl.user_session.get("history").append(
        {"role": "system", "content": "You are a helpful assistant."}
    )

    thread_id = cl.context.session.thread_id
    data_layer = get_data_layer()
    async with data_layer.engine.begin() as conn:
        await conn.execute(
            sql_text("INSERT INTO threads (id) VALUES (:tid) ON CONFLICT DO NOTHING"),
            {"tid": thread_id}
        )

    MODEL.reset_kv_cache()

    # ✅ Auto-ingest any files uploaded before chat start
    uploaded_files = cl.context.session.files
    if uploaded_files:
        for f in uploaded_files:
            try:
                await persist_file_content(f.path)
                print(f"✅ Auto-ingested {f.path} at chat start")
            except Exception as e:
                print(f"⚠️ Failed to ingest {f.path}: {e}")



def chunk_rows(rows, chunk_size=500):
    """Split large tabular datasets into manageable chunks."""
    for i in range(0, len(rows), chunk_size):
        yield rows[i:i+chunk_size]

def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 100):
    """
    Split raw text into overlapping chunks for FAISS indexing.
    - chunk_size: max characters per chunk
    - overlap: repeated characters between chunks for context continuity
    """
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        chunks.append(chunk)
        start += chunk_size - overlap
    return chunks

async def persist_file_content(file_path: str):
    """
    Unified file ingestion + persistence with file-type specific chunking:
    - Excel/CSV/JSON → chunk rows (default 500)
    - Text/PDF/DOCX → chunk text (default 1000 chars, 100 overlap)
    - Images → OCR text chunked like TXT
    """
    data_layer = get_data_layer()
    thread_id = cl.context.session.thread_id
    file_name = Path(file_path).name
    ext = Path(file_path).suffix.lower()

    # ✅ Use dispatcher to ingest file
    headers, rows_or_text = ingest_file(file_path)

    # Structured files (Excel, CSV, JSON)
    if ext in [".xlsx", ".xls", ".csv", ".json"] and headers and rows_or_text:
        rows = rows_or_text
        for chunk in chunk_rows(rows, chunk_size=500):   # 500 rows per chunk
            payload = {
                "thread_id": thread_id,
                "file_name": file_name,
                "headers": json.dumps(headers),
                "rows": json.dumps([sanitize_row(r) for r in chunk]),
                "text_content": None,
            }
            async with data_layer.engine.begin() as conn:
                await conn.execute(
                    sql_text("""
                        INSERT INTO file_contents (thread_id, file_name, headers, rows, text_content)
                        VALUES (:thread_id, :file_name, :headers, :rows, :text_content)
                    """),
                    payload
                )
            add_to_index(chunk, headers)

    # Unstructured text files (TXT, DOCX, PDF)
    else:
        text_content = rows_or_text
        payload = {
            "thread_id": thread_id,
            "file_name": file_name,
            "headers": None,
            "rows": None,
            "text_content": text_content if isinstance(text_content, str) else json.dumps(text_content),
        }
        async with data_layer.engine.begin() as conn:
            await conn.execute(
                sql_text("""
                    INSERT INTO file_contents (thread_id, file_name, headers, rows, text_content)
                    VALUES (:thread_id, :file_name, :headers, :rows, :text_content)
                """),
                payload
            )

        # ✅ Chunk text for FAISS indexing
        if isinstance(text_content, str):
            for chunk in chunk_text(text_content, chunk_size=1000, overlap=100):  # 1000 chars per chunk
                add_to_index([{"RawText": chunk}], ["RawText"])
        elif isinstance(text_content, list):
            for chunk in text_content:
                for sub_chunk in chunk_text(chunk, chunk_size=1000, overlap=100):
                    add_to_index([{"RawText": sub_chunk}], ["RawText"])

    print(f"✅ Saved {file_name} to Postgres and FAISS for thread {thread_id}")


def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = []
    for i, cell in enumerate(next(sheet.iter_rows(values_only=True))):
        headers.append(str(cell).strip() if cell else f"Column{i}")
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        rows.append(sanitize_row(row_dict))   # ✅ sanitize each row
    return headers, rows


def load_csv(file_path):
    rows, headers = [], []
    with open(file_path, newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        for row in reader:
            rows.append(sanitize_row(row))   # ✅ sanitize each row
    return headers, rows


def load_docx(file_path):
    """
    Try Docling first (Markdown).
    Fallback to python-docx if Docling fails or offline.
    """
    try:
        result = doc_converter.convert(file_path)
        return result.to_markdown()
    except Exception as e:
        print(f"⚠️ Docling failed, falling back: {e}")
        import docx
        doc = docx.Document(file_path)
        return [para.text for para in doc.paragraphs if para.text.strip()]


def load_pdf(file_path):
    """
    Try LlamaParse first (clean Markdown).
    Fallback to PyPDF2 + OCR if LlamaParse fails or offline.
    """
    try:
        return parser.parse(file_path)
    except Exception as e:
        print(f"⚠️ LlamaParse failed, falling back: {e}")
        texts = []
        reader = PyPDF2.PdfReader(file_path)
        POPPLER_PATH = r"C:\poppler-26.02.0\Library\bin"
        for page_num, page in enumerate(reader.pages):
            try:
                page_text = page.extract_text()
                if page_text and page_text.strip():
                    texts.append(page_text)
                else:
                    images = convert_from_path(
                        file_path,
                        first_page=page_num + 1,
                        last_page=page_num + 1,
                        poppler_path=POPPLER_PATH
                    )
                    for img in images:
                        ocr_text = pytesseract.image_to_string(img)
                        if ocr_text.strip():
                            texts.append(ocr_text)
            except Exception as e2:
                texts.append(f"⚠️ OCR error on page {page_num+1}: {e2}")
        return texts if texts else ["⚠️ No text extracted from PDF."]


    
def load_txt(file_path):
    """Load plain text files."""
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        file_text = f.read()
    return file_text.splitlines()


def load_json(file_path):
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        data = json.load(f)

    if isinstance(data, list) and all(isinstance(item, dict) for item in data):
        headers = list(data[0].keys()) if data else []
        rows = [sanitize_row(r) for r in data]   # ✅ sanitize JSON rows
        return headers, rows
    else:
        # fallback: treat as raw text
        file_text = json.dumps(data, indent=2, default=str)  # ✅ ensure safe serialization
        return [], [{"RawJSON": file_text}]

# Add Dispatcher Functionn

def ingest_file(file_path: str):
    """
    Universal file dispatcher:
    Auto-detects file type and routes to the correct loader.
    Returns (headers, rows) for structured files or (None, text_chunks) for unstructured.
    """
    ext = Path(file_path).suffix.lower()

    if ext in [".xlsx", ".xls"]:
        return load_excel(file_path)
    elif ext == ".csv":
        return load_csv(file_path)
    elif ext == ".docx":
        # Docx returns paragraphs, so wrap in rows
        return [], load_docx(file_path)
    elif ext == ".pdf":
        return [], load_pdf(file_path)
    elif ext == ".txt":
        return [], load_txt(file_path)
    elif ext == ".json":
        return load_json(file_path)
    else:
        # Fallback: treat as raw text
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return [], [{"RawText": f.read()}]

async def list_ingested_files():
    """
    List all ingested files for the current thread/session.
    Shows file name, headers, and number of rows/text chunks.
    """
    thread_id = cl.context.session.thread_id
    data_layer = get_data_layer()

    async with data_layer.engine.begin() as conn:
        result = await conn.execute(
            sql_text("SELECT file_name, headers, rows, text_content FROM file_contents WHERE thread_id = :tid"),
            {"tid": thread_id}
        )

        files_info = []
        for row in result:
            headers = json.loads(row.headers) if row.headers else []
            rows = json.loads(row.rows) if row.rows else []
            text_content = row.text_content

            files_info.append({
                "file_name": row.file_name,
                "headers": headers,
                "row_count": len(rows) if rows else 0,
                "text_length": len(text_content) if text_content else 0
            })

        if not files_info:
            await cl.Message(content="⚠️ No files ingested yet for this session.").send()
        else:
            msg = "📂 **Ingested Files in Current Session:**\n\n"
            for f in files_info:
                msg += f"- **{f['file_name']}**\n"
                if f["headers"]:
                    msg += f"  • Headers: {', '.join(f['headers'][:5])}...\n"
                msg += f"  • Rows: {f['row_count']}\n"
                msg += f"  • Text length: {f['text_length']} characters\n\n"
            await cl.Message(content=msg).send()


async def delete_ingested_file(file_name: str):
    """
    Delete an ingested file from the current session.
    Removes it from Postgres and clears it from FAISS index.
    """
    thread_id = cl.context.session.thread_id
    data_layer = get_data_layer()

    async with data_layer.engine.begin() as conn:
        result = await conn.execute(
            sql_text("DELETE FROM file_contents WHERE thread_id = :tid AND file_name = :fname"),
            {"tid": thread_id, "fname": file_name}
        )

    # ⚠️ Note: FAISS index cleanup depends on your add_to_index implementation.
    # If you want to fully remove vectors, you’ll need a delete_from_index helper.
    # For now, this removes the file from Postgres so it won’t be reloaded.

    await cl.Message(content=f"🗑️ File **{file_name}** has been deleted from this session.").send()



def auto_detect_column(headers, user_message):
    msg_lower = user_message.lower()
    for h in headers:
        if h.lower() in ["project name", "department"]:
            return h
    for h in headers:
        if h.lower() in msg_lower:
            return h
    return headers[0] if headers else None

from difflib import get_close_matches

def fuzzy_detect_column(headers, user_message, cutoff=0.6):
    """
    Fuzzy column detector for a single file:
    - Matches user query words to headers even if not exact.
    - Example: "Sale" -> "Sales", "Dept" -> "Department".
    Works with ANY file headers (Revenue, ClientName, Budget2026, etc.).
    """
    msg_lower = user_message.lower()
    # Try exact match first
    for h in headers:
        if h.lower() in msg_lower:
            return h

    # Fuzzy match using difflib
    for word in msg_lower.split():
        matches = get_close_matches(word, headers, n=1, cutoff=cutoff)
        if matches:
            return matches[0]

    # Fallback: first header if nothing matches
    return headers[0] if headers else None

def fuzzy_detect_column_multi(file_headers_dict, user_message, cutoff=0.6):
    """
    Fuzzy column detector across multiple files:
    - file_headers_dict: {filename: [headers]}
    - Returns a unified mapping of user query -> best header per file.
    Example: {"file1.xlsx": ["Dept"], "file2.csv": ["Department"]}
             query "department" -> {"file1.xlsx": "Dept", "file2.csv": "Department"}
    """
    unified_mapping = {}
    msg_lower = user_message.lower()

    for fname, headers in file_headers_dict.items():
        # Try exact match first
        matched = None
        for h in headers:
            if h.lower() in msg_lower:
                matched = h
                break

        # Fuzzy match if no exact match
        if not matched:
            for word in msg_lower.split():
                matches = get_close_matches(word, headers, n=1, cutoff=cutoff)
                if matches:
                    matched = matches[0]
                    break

        # Fallback: first header if nothing matches
        if not matched and headers:
            matched = headers[0]

        unified_mapping[fname] = matched

    return unified_mapping


def extract_keywords(user_message, headers, rows):
    msg_lower = user_message.lower()
    found = set()
    for h in headers:
        for r in rows:
            val = str(r.get(h, "")).strip().lower()
            if val and val in msg_lower:
                found.add(val)
    if found:
        return list(found)
    msg_parts = msg_lower.split()
    return [" ".join(msg_parts[-3:]).strip()]

def filter_rows_by_values(rows, column_name, values):
    results = []
    for r in rows:
        cell_val = str(r.get(column_name, "")).strip().lower()
        if any(v in cell_val for v in values):
            results.append(r)
    return results

def parse_numeric(value):
    """
    Robust universal numeric parser:
    - Handles strings like '$50,994', '€150,000', '12%', 'Age: 45', 'EUR 2,500', '₦1,200,000'
    - Strips currency symbols, commas, %, text noise
    - Returns float if parsing succeeds, else None
    """
    if value is None:
        return None

    raw_text = str(value).strip()

    # Remove common currency symbols and unit labels
    raw_text = re.sub(r"(USD|EUR|GBP|JPY|CFA|NGN|INR|CAD|AUD|CHF|₦|₿|¥|€|£|\$)", "", raw_text, flags=re.IGNORECASE)

    # Remove commas, percent signs, and extra words
    raw_text = re.sub(r"[,%]", "", raw_text)
    raw_text = re.sub(r"[A-Za-z:]", "", raw_text)  # strip stray letters like 'Age:' or 'Salary'

    # Keep only digits, dot, minus
    cleaned = re.sub(r"[^\d\.\-]", "", raw_text)

    try:
        return float(cleaned)
    except ValueError:
        return None

def filter_rows_by_numeric(rows, column_name, threshold, operator=">"):
    """
    Universal numeric filter for ANY file type:
    - rows: list of dicts (from Excel, CSV, JSON, PDF text, DOCX tables, etc.)
    - column_name: header to check (e.g., 'Annual Salary')
    - threshold: numeric threshold (e.g., 150000)
    - operator: comparison ('>', '<', '>=', '<=', '==')

    Returns only rows that satisfy the condition.
    """
    results = []
    for r in rows:
        val = parse_numeric(r.get(column_name))
        if val is None:
            continue

        if operator == ">" and val > threshold:
            results.append(r)
        elif operator == "<" and val < threshold:
            results.append(r)
        elif operator == ">=" and val >= threshold:
            results.append(r)
        elif operator == "<=" and val <= threshold:
            results.append(r)
        elif operator == "==" and val == threshold:
            results.append(r)

    return results

def get_filtered_table(rows, column_name, values, headers=None):
    """
    Filter rows by specific values in a column and return a Markdown table.
    """
    filtered = filter_rows_by_values(rows, column_name, values)
    if not headers and filtered:
        headers = list(filtered[0].keys())
    if not filtered:
        return "No results found."
    # Return Markdown table
    return build_markdown_table(filtered, headers)

def build_table(rows, headers, limit=20):
    """
    Build a Markdown table from rows and headers.
    """
    if not rows:
        return "No matching records found."
    # Return Markdown table
    return build_markdown_table(rows[:limit], headers)

def build_markdown_table(rows, headers, limit=None):
    """
    Render rows as a Markdown table with proper headers.
    """
    if not rows:
        return "No matching records found."
    if limit:
        rows = rows[:limit]

    # Build header row
    table = "| " + " | ".join(headers) + " |\n"
    table += "| " + " | ".join(["---"] * len(headers)) + " |\n"

    # Build data rows
    for r in rows:
        row_values = [str(r.get(h, "")) for h in headers]
        table += "| " + " | ".join(row_values) + " |\n"

    return table


def compute_all_numeric_stats(rows: list, headers: list) -> str:
    """
    Auto-detect all numeric columns in any file table (Excel, CSV, JSON, etc.)
    and compute sum, average, min, max for each.
    """
    numeric_stats = {}

    for col in headers:
        values = []
        for r in rows:
            val = r.get(col)
            try:
                num = float(val)
                values.append(num)
            except (TypeError, ValueError):
                continue

        if values:
            numeric_stats[col] = {
                "count": len(values),
                "sum": sum(values),
                "avg": sum(values) / len(values),
                "min": min(values),
                "max": max(values),
            }

    if not numeric_stats:
        return "⚠️ No numeric columns found to compute stats."

    # Build Markdown summary
    result = "📊 **Numeric Stats for All Columns:**\n\n"
    for col, stats in numeric_stats.items():
        result += (
            f"- Column **{col}**:\n"
            f"  • Count: {stats['count']}\n"
            f"  • Sum: {stats['sum']:.2f}\n"
            f"  • Average: {stats['avg']:.2f}\n"
            f"  • Min: {stats['min']:.2f}\n"
            f"  • Max: {stats['max']:.2f}\n\n"
        )
    return result

import re

def detect_conditions(user_message: str, headers: list):
    """
    Detects numeric and string conditions from user query.
    Supports AND / OR logic.
    Returns a list of (column_name, value/threshold, operator, logic).
    """

    msg_lower = user_message.lower()

    # Map words to operators
    operator_map = {
        "above": ">",
        "greater than": ">",
        "over": ">",
        "below": "<",
        "less than": "<",
        "under": "<",
        "at least": ">=",
        "minimum": ">=",
        "not less than": ">=",
        "at most": "<=",
        "maximum": "<=",
        "not more than": "<=",
        "equal to": "==",
        "equals": "==",
        "exactly": "==",
        "=": "=="
    }

    conditions = []
    parts = re.split(r"\b(and|or)\b", msg_lower)

    logic = "AND"  # default
    for part in parts:
        part = part.strip()
        if part == "and":
            logic = "AND"
            continue
        elif part == "or":
            logic = "OR"
            continue

        # Detect operator
        operator = None
        for phrase, op in operator_map.items():
            if phrase in part:
                operator = op
                break

        # Detect numeric threshold
        match = re.search(r"(\d[\d,\.]*)", part)
        threshold = None
        if match:
            threshold = float(match.group(1).replace(",", ""))

        # Detect column name by fuzzy match
        column_name = None
        for h in headers:
            if h.lower() in part:
                column_name = h
                break

        # String condition (e.g., Department = Finance)
        if operator == "==" and not threshold:
            # Extract word after '=' or 'equals'
            str_match = re.search(r"(?:=|equals)\s+([a-zA-Z0-9\s]+)", part)
            if str_match and column_name:
                value = str_match.group(1).strip()
                conditions.append((column_name, value, "==", logic))
                continue

        # Numeric condition
        if operator and threshold and column_name:
            conditions.append((column_name, threshold, operator, logic))

    return conditions


def apply_conditions(rows: list, conditions: list):
    """
    Apply multiple numeric and string conditions with AND/OR logic.
    """
    if not conditions:
        return rows

    # Start with first condition
    col, val, op, _ = conditions[0]
    if isinstance(val, (int, float)):
        filtered = filter_rows_by_numeric(rows, col, val, operator=op)
    else:
        filtered = filter_rows_by_values(rows, col, [val])

    for col, val, op, logic in conditions[1:]:
        if isinstance(val, (int, float)):
            next_filtered = filter_rows_by_numeric(rows, col, val, operator=op)
        else:
            next_filtered = filter_rows_by_values(rows, col, [val])

        if logic == "AND":
            # Intersection
            filtered = [r for r in filtered if r in next_filtered]
        elif logic == "OR":
            # Union
            filtered = list({id(r): r for r in (filtered + next_filtered)}.values())

    return filtered

async def debug_file_ingestion(file_name: str, file_table: list, headers: list, file_text: str):
    """
    Diagnostic helper: prints ingestion status for ANY file type.
    Shows whether structured rows (file_table) or raw text (file_text) are available.
    """

    print("\n📂 Debugging File Ingestion")
    print(f"➡️ File: {file_name}")

    if headers:
        print(f"📝 Headers detected: {headers}")
    else:
        print("⚠️ No headers detected.")

    if file_table:
        print(f"✅ Structured rows loaded: {len(file_table)} rows")
        for r in file_table[:2]:
            print(f"   Row preview: {r}")
    else:
        print("⚠️ No structured rows found.")

    if file_text and file_text.strip():
        print(f"✅ Raw text extracted: {len(file_text)} characters")
        print(f"   Text preview: {file_text[:200]}...")
    else:
        print("⚠️ No raw text extracted.")

    print("📊 Ingestion check complete.\n")


def auto_detect_columns(headers, rows):
    categorical_cols, numeric_cols, date_cols = [], [], []
    for h in headers:
        values = [str(r.get(h, "")) for r in rows if r.get(h)]
        if not values:
            continue
        numeric_count, date_count = 0, 0
        for v in values[:20]:
            try:
                float(v); numeric_count += 1
            except:
                try:
                    datetime.datetime.fromisoformat(v); date_count += 1
                except:
                    pass
        if numeric_count > len(values)*0.5:
            numeric_cols.append(h)
        elif date_count > len(values)*0.3:
            date_cols.append(h)
        else:
            categorical_cols.append(h)
    return categorical_cols, numeric_cols, date_cols

def build_copy_table(rows, headers, limit=30):
    if not rows:
        return "No matching records found."
    # Plain text tab-separated table
    table = "\t".join(headers) + "\n"
    for r in rows[:limit]:
        row_values = [str(r.get(h, "")) for h in headers]
        table += "\t".join(row_values) + "\n"
    return table

def compute_all_numeric_stats(rows, headers):
    result = "Numeric Stats:\n"
    found = False
    for col in headers:
        values = []
        for r in rows:
            try:
                values.append(float(r.get(col)))
            except:
                continue
        if values:
            found = True
            result += f"{col}: Count={len(values)}, Sum={sum(values):.2f}, Avg={sum(values)/len(values):.2f}, Min={min(values):.2f}, Max={max(values):.2f}\n"
    return result if found else "No numeric columns found."

def detect_trends(rows, numeric_cols, date_cols):
    """
    Sorts by first date column and compares numeric values over time.
    Returns a plain text summary of trends.
    """
    if not numeric_cols or not date_cols:
        return ""

    date_col = date_cols[0]
    trends = "Trends:\n"
    try:
        sorted_rows = sorted(
            rows,
            key=lambda r: datetime.datetime.fromisoformat(str(r.get(date_col)))
            if r.get(date_col) else datetime.datetime.min
        )
        for num_col in numeric_cols:
            values = [(r.get(date_col), parse_numeric(r.get(num_col))) for r in sorted_rows if parse_numeric(r.get(num_col)) is not None]
            if len(values) >= 2:
                start_date, start_val = values[0]
                end_date, end_val = values[-1]
                direction = "increased" if end_val > start_val else "decreased" if end_val < start_val else "remained stable"
                trends += f"{num_col}: {direction} from {start_val} → {end_val} between {start_date} and {end_date}\n"
    except Exception as e:
        trends += f"Trend detection error: {e}\n"

    return trends


async def professional_llm_response(user_query, file_table, headers, llm_runner, file_text="", preview_limit=30):
    """
    Context-managed universal response:
    - Queries FAISS first
    - Falls back to raw file content if FAISS returns nothing
    - Handles chunking, stats, trends, enrichment
    """

    MODEL.reset_kv_cache()

    # ✅ Query FAISS first
    relevant_rows = query_index(user_query, top_k=10)
    filtered_rows = relevant_rows if relevant_rows else (file_table or [])

    # ✅ If FAISS returns nothing and no rows, fallback to raw text (chunked)
    if not filtered_rows and file_text:
        if isinstance(file_text, str):
            filtered_rows = [{"RawText": chunk} for chunk in chunk_text(file_text)]
        elif isinstance(file_text, list):
            filtered_rows = []
            for chunk in file_text:
                filtered_rows.extend([{"RawText": sub_chunk} for sub_chunk in chunk_text(chunk)])

    # ✅ Apply conditions if headers exist
    conditions = detect_conditions(user_query, headers) if headers else []
    if conditions and file_table:
        filtered_rows = apply_conditions(file_table, conditions) or []

    # ✅ Trim oversized datasets
    MAX_TOTAL_ROWS = 2000
    if len(filtered_rows) > MAX_TOTAL_ROWS:
        filtered_rows = filtered_rows[:MAX_TOTAL_ROWS]

    # ✅ Chunk rows
    chunked_results = []
    for chunk in chunk_rows(filtered_rows, 200):
        if headers:
            json_summary = [{h: row.get(h, "") for h in headers} for row in chunk]
            categorical_cols, numeric_cols, date_cols = auto_detect_columns(headers, chunk)
            numeric_stats = compute_all_numeric_stats(chunk, headers) if numeric_cols else ""
            table_text = build_copy_table(chunk, headers, preview_limit) if chunk else ""
            trend_summary = detect_trends(chunk, numeric_cols, date_cols)
        else:
            json_summary = chunk
            numeric_stats, table_text, trend_summary = "", "", ""

        chunk_prompt = f"User question: {user_query}\n\nChunk size: {len(chunk)} rows.\n\nJSON Summary:\n{json.dumps(json_summary, indent=2)}\n"
        if numeric_stats: chunk_prompt += "\n" + numeric_stats
        if table_text: chunk_prompt += "\nFiltered Results:\n" + table_text
        if trend_summary: chunk_prompt += "\n" + trend_summary

        chunk_answer = await llm_runner(chunk_prompt)
        chunked_results.append(chunk_answer)

    # ✅ Domain enrichment
    enrichment = ""
    blob = " ".join(headers).lower() + " " + user_query.lower() + " " + str(file_text).lower()
    if any(k in blob for k in ["finance","revenue","expenses","budget"]):
        enrichment = "Accounting Insight → Expenses rising faster than revenue."
    elif any(k in blob for k in ["hr","employee","staff","training"]):
        enrichment = "HR Insight → Monitor turnover and training."
    elif "sales" in blob:
        enrichment = "Sales Insight → Track conversion rates."
    elif "project" in blob or "task" in blob:
        enrichment = "Project Insight → Monitor deadlines and milestones."
    elif "training" in blob or "course" in blob or "education" in blob:
        enrichment = "Education Insight → Average grade trends detected."
    elif "medical" in blob or "patient" in blob or "blood pressure" in blob:
        enrichment = "Medical Insight → High blood pressure risk detected."
    elif "logistics" in blob or "delivery" in blob:
        enrichment = "Logistics Insight → Delivery times increasing, check supply chain."
    elif "engineering" in blob or "resource" in blob:
        enrichment = "Engineering Insight → Resource utilization efficiency detected."
    elif "research" in blob or "publication" in blob:
        enrichment = "Research Insight → Publication output trends identified."

    final_answer = "\n\n".join(chunked_results)
    if enrichment:
        final_answer += "\n" + enrichment

    return final_answer



async def translate_text(text: str, target_lang: str) -> str:
    """
    Simple translation helper using Qwen3 itself.
    """
    prompt = f"Translate the following text into {target_lang}:\n\n{text}"
    return await cl.make_async(run_qwen_sync)(prompt)

# ✅ Insert here
async def summarize_chunk(chunk: str, lang: str) -> str:
    """Summarize a single chunk asynchronously."""
    prompt = f"Summarize this chunk in {lang}:\n{chunk}"
    summary = await cl.make_async(run_qwen_sync)(prompt)
    return f"🔹 {summary}"

async def summarize_file_text(file_text: str, lang: str = "English") -> str:
    """
    Parallel summarization for large text files.
    Splits text into chunks and runs all summaries concurrently.
    """
    chunk_size = 4000
    chunks = [file_text[i:i+chunk_size] for i in range(0, len(file_text), chunk_size)]

    # ✅ Run all summarizations in parallel
    tasks = [summarize_chunk(chunk, lang) for chunk in chunks[:10]]
    summaries = await asyncio.gather(*tasks)

    return f"📄 **Summary (in {lang}):**\n\n" + "\n".join(summaries)

async def handle_multi_file_query(user_query: str, files: list, use_multi_fuzzy: bool = True) -> str:
    """
    Cross-file reasoning handler with optional multi-file fuzzy detection:
    - Merges tables and text from multiple files
    - Runs numeric stats for structured data
    - Summarizes unstructured text
    - Applies domain-specific enrichment automatically
    - Optionally unifies similar headers across files (Dept vs Department)
    """

    merged_tables = []
    merged_headers = set()
    merged_texts = []
    file_headers_dict = {}

    # Collect data from all files
    for f in files:
        headers = cl.user_session.get(f"{f}_headers", [])
        rows = cl.user_session.get(f"{f}_table", [])
        text = cl.user_session.get(f"{f}_text", "")

        if headers:
            file_headers_dict[f] = headers
        if rows:
            merged_tables.extend(rows)
            merged_headers.update(headers)
        if text:
            merged_texts.append(text)

    # 🔹 Fuzzy detection (optional)
    unified_cols = {}
    if use_multi_fuzzy and file_headers_dict:
        unified_cols = fuzzy_detect_column_multi(file_headers_dict, user_query)
    else:
        if merged_headers:
            unified_cols = {files[0]: fuzzy_detect_column(list(merged_headers), user_query)}

    # 🔹 Numeric stats across all files
    stats_summary = ""
    if merged_tables and merged_headers:
        stats_summary = compute_all_numeric_stats(merged_tables, list(merged_headers))

    # 🔹 Summarize text across all files
    text_summary = ""
    if merged_texts:
        combined_text = "\n".join(merged_texts)
        text_summary = await summarize_file_text(combined_text, lang="English")

    # 🔹 Domain-specific enrichment
    enrichment = ""
    st_model = cl.user_session.get("st_model")
    for f in files:
        texts = cl.user_session.get(f"{f}_text", "").split("\n")
        if texts and st_model:
            domain = await detect_file_domain(texts, st_model)
            if domain == "Finance":
                enrichment += "\n💰 **Finance Insights:**\n- Profit margin analysis\n- Growth trends\n"
            elif domain == "Legal":
                enrichment += "\n⚖️ **Legal Insights:**\n- Clause extraction\n- Compliance risks\n"
            elif domain == "HR":
                enrichment += "\n👥 **HR Insights:**\n- Salary distribution\n- Turnover detection\n"
            elif domain == "Project":
                enrichment += "\n📈 **Project Insights:**\n- Milestone tracking\n- Deadline risks\n"
            elif domain == "Compliance":
                enrichment += "\n🛡️ **Compliance Insights:**\n- Policy adherence\n- Risk factors\n"
            elif domain == "Customer":
                enrichment += "\n🤝 **Customer Insights:**\n- Order trends\n- Feedback themes\n"

    # 🔹 Build unified answer
    final_answer = "📂 **Cross‑File Unified Answer:**\n\n"
    if unified_cols:
        final_answer += f"🔎 **Unified Column Mapping:** {unified_cols}\n\n"
    if stats_summary:
        final_answer += stats_summary + "\n"
    if text_summary:
        final_answer += text_summary + "\n"
    if enrichment:
        final_answer += enrichment + "\n"

    if not stats_summary and not text_summary and not enrichment:
        final_answer += "⚠️ No usable data found across files."

    return final_answer

async def detect_file_domain(texts: list, st_model) -> str:
    """
    Detects the domain of a file using embeddings + similarity classification.
    Domains: Finance, Legal, HR, Project, Compliance, Customer
    """

    # Candidate domain labels
    domains = {
        "Finance": ["revenue", "sales", "profit", "expenses", "budget"],
        "Legal": ["contract", "agreement", "clause", "obligation", "law"],
        "HR": ["employee", "salary", "turnover", "department", "hiring"],
        "Project": ["project", "milestone", "deadline", "progress", "task"],
        "Compliance": ["policy", "audit", "compliance", "risk", "regulation"],
        "Customer": ["customer", "client", "order", "feedback", "retention"]
    }

    # Encode sample text
    sample = " ".join(texts[:5])  # take first few chunks
    sample_vec = st_model.encode([sample])

    # Encode domain keywords
    scores = {}
    for domain, keywords in domains.items():
        keyword_vec = st_model.encode([" ".join(keywords)])
        similarity = float(np.dot(sample_vec, keyword_vec.T) / (np.linalg.norm(sample_vec) * np.linalg.norm(keyword_vec)))
        scores[domain] = similarity

    # Pick best match
    best_domain = max(scores, key=scores.get)
    return best_domain

async def debug_file_ingestion(file_name: str, file_table: list, headers: list, file_text: str):
    """
    Diagnostic helper: prints ingestion status for ANY file type.
    Shows whether structured rows (file_table) or raw text (file_text) are available.
    """

    print("\n📂 Debugging File Ingestion")
    print(f"➡️ File: {file_name}")

    if headers:
        print(f"📝 Headers detected: {headers}")
    else:
        print("⚠️ No headers detected.")

    if file_table:
        print(f"✅ Structured rows loaded: {len(file_table)} rows")
        for r in file_table[:2]:
            print(f"   Row preview: {r}")
    else:
        print("⚠️ No structured rows found.")

    if file_text and file_text.strip():
        print(f"✅ Raw text extracted: {len(file_text)} characters")
        print(f"   Text preview: {file_text[:200]}...")
    else:
        print("⚠️ No raw text extracted.")

    print("📊 Ingestion check complete.\n")


async def handle_user_query(
    msg_content: str,
    file_table: list,
    headers: list,
    file_text: str,
    file_path: str
):
    """
    Unified query handler:
    - Uses actual Excel headers (no ColumnX fallback)
    - Applies numeric/string filters
    - Builds a clean Representative Table with build_table()
    - Drops language detection completely
    """

    # ✅ Diagnostic check: print ingestion status
    await debug_file_ingestion(file_path, file_table, headers, file_text)


    # If tabular data is available → structured response
    if file_table and headers:
        # Apply conditions
        conditions = detect_conditions(msg_content, headers)
        filtered_rows = apply_conditions(file_table, conditions) if conditions else file_table

        # Build table
        table = build_table(filtered_rows, headers, limit=20)

        # Numeric stats
        numeric_stats = compute_all_numeric_stats(filtered_rows, headers) if filtered_rows else ""

        # Return raw table + stats (no enforced format)
        return f"{table}\n\n{numeric_stats}" if (table or numeric_stats) else "⚠️ No matching records found."

    # Fallback: semantic search
    index = cl.user_session.get("file_index")
    st_model = cl.user_session.get("st_model")
    texts = cl.user_session.get("file_text", "").split("\n")

    if index and st_model and texts:
        query_vec = st_model.encode([msg_content])
        D, I = index.search(np.array(query_vec), k=3)
        if len(I[0]) > 0:
            context = "\n".join([texts[i] for i in I[0]])
            prompt = f"User question: {msg_content}\n\nRelevant context:\n{context}\n\nAnswer clearly."
            llm_answer = await cl.make_async(run_qwen_sync)(prompt)
            return llm_answer

    # Fallback: summarization
    if file_text:
        return await summarize_file_text(file_text, "English")

    return "⚠️ No data available to answer your query."


async def build_faiss_index(texts):
    word_embedding_model = modules.Transformer("all-MiniLM-L6-v2")
    pooling_model = modules.Pooling(word_embedding_model.get_embedding_dimension())
    st_model = SentenceTransformer(modules=[word_embedding_model, pooling_model])

    batch_size = 200
    total_batches = (len(texts) + batch_size - 1) // batch_size

    embeddings_list = []
    for i in range(total_batches):
        batch = texts[i*batch_size:(i+1)*batch_size]

        # ✅ Console log only
        print(f"Processing batch {i+1}/{total_batches} in PowerShell...")

        # Run asynchronously to avoid blocking
        batch_embeddings = await cl.make_async(st_model.encode)(batch, convert_to_numpy=True)
        embeddings_list.append(batch_embeddings)

    embeddings = np.vstack(embeddings_list)
    dimension = embeddings.shape[1]
    index = faiss.IndexFlatL2(dimension)
    index.add(embeddings)

    return index, st_model

async def prepare_file_context(file_path: str):
    """
    Clean ingestion helper:
    - Handles Excel, CSV, DOCX, PDF, TXT, JSON, and images.
    - Builds FAISS index for semantic search.
    - No language detection.
    """

    MODEL.reset_kv_cache()
    file_path = os.path.abspath(file_path)
    ext = os.path.splitext(file_path)[1].lower()
    headers, rows, texts = [], [], []

    # File type handling
    if ext in [".xls", ".xlsx"]:
        headers, rows = load_excel(file_path)
    elif ext == ".csv":
        headers, rows = load_csv(file_path)
    elif ext == ".docx":
        texts = load_docx(file_path)
    elif ext == ".pdf":
        texts = load_pdf(file_path)
    elif ext == ".txt":
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            texts = f.readlines()
    elif ext in [".png", ".jpg", ".jpeg"]:
        img = PILImage.open(file_path)
        texts = [pytesseract.image_to_string(img)]
    elif ext == ".json":
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        texts = [json.dumps(data, indent=2)]
    else:
        texts = [f"Unsupported file type: {ext}"]

    # Chunk rows for tabular files
    if rows:
        texts = []
        for chunk in chunk_rows(rows):
            texts.extend([
                " | ".join(f"{k}: {str(v)}" for k, v in r.items() if v not in [None, ""])
                for r in chunk if any(r.values())
            ])
        cl.user_session.set("file_table", rows)
        cl.user_session.set("file_headers", headers)

    # ✅ Run FAISS indexing only (no language detection)
    if texts:
        index, st_model = await build_faiss_index(texts)

        cl.user_session.set("file_index", index)
        cl.user_session.set("file_text", "\n".join(texts))
        cl.user_session.set("st_model", st_model)

        return index, texts, st_model

    return None, texts, None

@cl.on_message
async def on_message(msg: cl.Message):
    """
    Unified handler with context management:
    - Maintains conversation history (dicts + HumanMessage/AIMessage objects)
    - Handles file ingestion (Excel, CSV, DOCX, PDF, TXT, JSON, Images with OCR)
    - Builds sidebar previews (Markdown table + numeric stats)
    - Saves ingested content into Postgres
    - Reloads file context safely (type‑checked JSON decode)
    - Uses professional_llm_response with chunking for large files
    """

    # ✅ Maintain conversation state (LLM objects)
    state = cl.user_session.get("state") or {"messages": []}
    cl.user_session.set("state", state)
    state["messages"].append(HumanMessage(content=msg.content))

    # ✅ Maintain conversation history (dicts)
    history = cl.user_session.get("history") or []
    history.append({"role": "user", "content": msg.content})
    cl.user_session.set("history", history)

    sidebar_elements = []

    # ✅ File ingestion inline
    if msg.elements:
        for element in msg.elements:
            if element.type == "file":
                file_path, file_name = element.path, element.name
                headers, rows, text_content = [], [], ""

                try:
                    if file_name.endswith((".xlsx", ".xls")):
                        headers, rows = load_excel(file_path)
                        preview = build_markdown_table(rows, headers, limit=5)
                        sidebar_elements.append(cl.Text(content=preview, name=f"{file_name} Preview"))
                        stats = compute_all_numeric_stats(rows, headers)
                        sidebar_elements.append(cl.Text(content=stats, name=f"{file_name} Stats"))
                        for chunk in chunk_rows(rows, 500):
                            await persist_file_content(file_path)

                    elif file_name.endswith(".csv"):
                        headers, rows = load_csv(file_path)
                        preview = build_markdown_table(rows, headers, limit=5)
                        sidebar_elements.append(cl.Text(content=preview, name=f"{file_name} Preview"))
                        stats = compute_all_numeric_stats(rows, headers)
                        sidebar_elements.append(cl.Text(content=stats, name=f"{file_name} Stats"))
                        for chunk in chunk_rows(rows, 500):
                            await persist_file_content(file_path)

                    elif file_name.endswith(".docx"):
                        texts = load_docx(file_path)
                        text_content = "\n".join(texts)
                        preview = "\n".join(texts[:10])
                        sidebar_elements.append(cl.Text(content=preview, name=file_name))
                        await persist_file_content(file_name, None, None, text_content)

                    elif file_name.endswith(".pdf"):
                        texts = load_pdf(file_path)
                        text_content = "\n".join(texts)
                        sidebar_elements.append(cl.Pdf(path=file_path, name=file_name))
                        await persist_file_content(file_name, None, None, text_content)

                    elif file_name.endswith(".txt"):
                        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                            text_content = f.read()
                        sidebar_elements.append(cl.Text(content=text_content[:500], name=file_name))
                        await persist_file_content(file_name, None, None, text_content)

                    elif file_name.endswith(".json"):
                        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                            data = json.load(f)
                        if isinstance(data, list) and all(isinstance(item, dict) for item in data):
                            headers = list(data[0].keys()) if data else []
                            preview = build_markdown_table(data, headers, limit=5)
                            sidebar_elements.append(cl.Text(content=preview, name=f"{file_name} Preview"))
                            stats = compute_all_numeric_stats(data, headers)
                            sidebar_elements.append(cl.Text(content=stats, name=f"{file_name} Stats"))
                            for chunk in chunk_rows(data, 500):
                                await persist_file_content(file_path)

                        else:
                            text_content = json.dumps(data, indent=2, default=str)
                            sidebar_elements.append(cl.Text(content=text_content[:500], name=file_name))
                            await persist_file_content(file_name, None, None, text_content)

                    elif file_name.lower().endswith((".png", ".jpg", ".jpeg")):
                        img = PILImage.open(file_path)
                        ocr_text = pytesseract.image_to_string(img)
                        if ocr_text.strip():
                            sidebar_elements.append(cl.Text(content=ocr_text[:500], name=file_name))
                            await persist_file_content(file_name, None, None, ocr_text)
                        sidebar_elements.append(cl.Image(path=file_path, name=file_name))

                    cl.user_session.set("last_file_name", file_name)
                    cl.user_session.set(f"{file_name}_headers", headers)
                    cl.user_session.set(f"{file_name}_table", rows)
                    cl.user_session.set(f"{file_name}_text", text_content)

                    print(f"✅ Ingested {file_name}")

                except Exception as e:
                    print(f"⚠️ Error reading file {file_name}: {e}")

    # ✅ Show sidebar previews
    if sidebar_elements:
        await cl.ElementSidebar.set_elements(sidebar_elements)
        await cl.ElementSidebar.set_title("Uploaded Files")

    # ✅ Collect file context from Postgres
    data_layer = get_data_layer()
    file_context = ""
    headers, rows, text_content = [], [], ""
    async with data_layer.engine.begin() as conn:
        result = await conn.execute(
            sql_text("SELECT file_name, headers, rows, text_content FROM file_contents WHERE thread_id = :tid"),
            {"tid": cl.context.session.thread_id}
        )
        for row in result:
            headers = row.headers
            rows = row.rows
            text_content = row.text_content if row.text_content else ""

            if isinstance(headers, str):
                headers = json.loads(headers)
            if isinstance(rows, str):
                rows = json.loads(rows)

            file_context += f"\n📂 File: {row.file_name}\n"
            if headers: file_context += f"Headers: {headers}\n"
            if rows: file_context += f"Rows: {rows[:2]}...\n"
            if text_content: file_context += f"Text: {text_content[:200]}...\n"

    # ✅ Pagination logic
    preview_limit = 30
    page = 0
    if "next" in msg.content.lower():
        page = cl.user_session.get("page", 0) + 1
    elif "previous" in msg.content.lower():
        page = max(cl.user_session.get("page", 0) - 1, 0)
    cl.user_session.set("page", page)

    # ✅ Use chunked professional_llm_response for large files
    answer = await professional_llm_response(
        user_query=msg.content,
        file_table=rows,
        headers=headers,
        llm_runner=cl.make_async(run_qwen_sync),  # ✅ non-blocking
        file_text=file_context,
        preview_limit=preview_limit
    )

    # ✅ Save assistant reply into both state + history
    state["messages"].append(AIMessage(content=answer))
    history.append({"role": "assistant", "content": answer})
    cl.user_session.set("state", {"messages": state["messages"]})
    cl.user_session.set("history", history)

    # ✅ Send answer back
    await cl.Message(content=answer).send()



@cl.on_stop
def on_stop():
    print("The user wants to stop the task!")

@cl.on_chat_end
def on_chat_end():
    print("The user disconnected!")

# ------------------- Starters -------------------
@cl.set_starters
async def set_starters():
    return [
        cl.Starter(
            label="Morning routine ideation",
            message="Can you help me create a personalized morning routine that would help increase my productivity throughout the day? Start by asking me about my current habits and what activities energize me in the morning.",
            icon="/public/idea.png"
        ),
        cl.Starter(
            label="Explain superconductors",
            message="Explain superconductors like I'm five years old.",
            icon="/public/learn.png"
        ),
        cl.Starter(
            label="Python script for daily email reports",
            message="Write a script to automate sending daily email reports in Python, and walk me through how I would set it up.",
            icon="/public/terminal.png",
            command="code"
        ),
        cl.Starter(
            label="Text inviting friend to wedding",
            message="Write a text asking a friend to be my plus-one at a wedding next month. I want to keep it super short and casual, and offer an out.",
            icon="/public/write.png"
        )
    ]

# Register tools based on environment flag
register_tools(cl)

