import os
import torch
import asyncio
import chainlit as cl
from typing import TypedDict, List
from pathlib import Path
from chainlit.types import ThreadDict
from chainlit.data.sql_alchemy import SQLAlchemyDataLayer
from langchain_core.messages import HumanMessage, BaseMessage, AIMessage
from langchain_community.document_loaders import PDFPlumberLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
import PyPDF2, docx, openpyxl
from PIL import Image as PILImage
import pytesseract

# Semantic search + embeddings
import faiss
from sentence_transformers import SentenceTransformer, models
import pandas as pd
import numpy as np

# Helpers for ingestion control
from helpers import request_cancel, request_pause, request_resume

# MCP + SocketIO
from mcp import ClientSession   # ✅ MCP integration
import socketio                 # ✅ Payload limit

from reasoning_from_scratch.ch02 import get_device, generate_text_basic_stream_cache
from reasoning_from_scratch.ch03 import load_model_and_tokenizer


# ============================================================
# Configuration
# ============================================================
WHICH_MODEL = "reasoning"
MAX_NEW_TOKENS = 4096
LOCAL_DIR = "qwen3"
CHECKPOINT_PATH = os.getenv("CHECKPOINT_PATH")
COMPILE = False
DEVICE = get_device()

def load_app_model_and_tokenizer():
    if CHECKPOINT_PATH is None:
        # Default: load from local dir
        return load_model_and_tokenizer(
            which_model=WHICH_MODEL,
            device=DEVICE,
            use_compile=COMPILE,
            local_dir=LOCAL_DIR,
        )
    checkpoint_path = Path(CHECKPOINT_PATH)
    if not checkpoint_path.exists():
        raise FileNotFoundError(f"Checkpoint file not found: {checkpoint_path}")
    # Load tokenizer only
    from reasoning_from_scratch.ch03 import load_tokenizer_only
    from reasoning_from_scratch.qwen3 import Qwen3Model, QWEN_CONFIG_06_B
    tokenizer = load_tokenizer_only(which_model=WHICH_MODEL, local_dir=LOCAL_DIR)
    model = Qwen3Model(QWEN_CONFIG_06_B)
    model.load_state_dict(torch.load(checkpoint_path, map_location="cpu"))
    model.to(DEVICE)
    if COMPILE:
        torch._dynamo.config.allow_unspec_int_on_nn_module = True
        model = torch.compile(model)
    return model, tokenizer

MODEL, TOKENIZER = load_app_model_and_tokenizer()

EOS_TOKEN_IDS = (
    TOKENIZER.encode("<|im_end|>")[0],
    TOKENIZER.encode("<|endoftext|>")[0]
)


import csv
import docx
from PyPDF2 import PdfReader
import openpyxl
import faiss
import numpy as np
from sentence_transformers import SentenceTransformer
from sentence_transformers.sentence_transformer import modules
import json
from PIL import Image as PILImage
import pytesseract

# ------------------ Header Detection --------------------
def detect_headers(sheet, max_scan=10):
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
        if len(non_empty) >= 2:
            headers = [str(cell).strip() if cell else f"Column{j}" for j, cell in enumerate(row)]
            return i, headers
    return 1, []

# ------------------ Loaders --------------------
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    header_row_index, headers = detect_headers(sheet)
    rows = []
    for row in sheet.iter_rows(min_row=header_row_index+1, values_only=True):
        if not any(row):
            continue
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        rows.append(row_dict)
    return headers, rows

def load_csv(file_path):
    rows, headers = [], []
    with open(file_path, newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        for row in reader:
            rows.append(row)
    return headers, rows

def load_docx(file_path):
    doc = docx.Document(file_path)
    texts = [para.text for para in doc.paragraphs if para.text.strip()]
    return texts

def load_pdf(file_path):
    reader = PdfReader(file_path)
    texts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            texts.append(text)
    return texts

# ------------------ Universal ingestion helper --------------------
async def prepare_file_context(file_path: str):
    """
    Universal ingestion helper:
    - Handles Excel, CSV, DOCX, PDF, TXT, JSON, and images.
    - Normalizes all values to strings.
    - Skips empty/malformed rows.
    - Wraps FAISS indexing in try/except to prevent crashes.
    """

    file_path = os.path.abspath(file_path)
    ext = os.path.splitext(file_path)[1].lower()

    headers, rows, texts = [], [], []

    try:
        if ext in [".xls", ".xlsx"]:
            headers, rows = load_excel(file_path)
            texts = [
                " | ".join(f"{k}: {str(v)}" for k, v in r.items() if v not in [None, ""])
                for r in rows if any(r.values())
            ]

        elif ext == ".csv":
            headers, rows = load_csv(file_path)
            texts = [
                " | ".join(f"{k}: {str(v)}" for k, v in r.items() if v not in [None, ""])
                for r in rows if any(r.values())
            ]

        elif ext == ".docx":
            texts = load_docx(file_path)

        elif ext == ".pdf":
            texts = load_pdf(file_path)

        elif ext == ".txt":
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                texts = f.readlines()

        elif ext in [".png", ".jpg", ".jpeg"]:
            img = PILImage.open(file_path)
            texts = [pytesseract.image_to_string(img)]

        elif ext == ".json":
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            texts = [json.dumps(data, indent=2)]

        else:
            texts = [f"Unsupported file type: {ext}"]

    except Exception as e:
        print(f"⚠️ Error ingesting {file_path}: {e}")
        texts = [f"Error reading file: {e}"]

    # ✅ Normalize texts: ensure all are strings and non-empty
    texts = [str(t).strip() for t in texts if str(t).strip()]

    # ✅ Safe FAISS indexing
    index, st_model = None, None
    if texts:
        try:
            word_embedding_model = modules.Transformer("all-MiniLM-L6-v2")
            pooling_model = modules.Pooling(word_embedding_model.get_embedding_dimension())
            st_model = SentenceTransformer(modules=[word_embedding_model, pooling_model])

            embeddings = st_model.encode(texts, convert_to_numpy=True)
            if embeddings.ndim == 2 and embeddings.shape[0] > 0:
                dimension = embeddings.shape[1]
                index = faiss.IndexFlatL2(dimension)
                index.add(embeddings)
        except Exception as e:
            print(f"⚠️ Embedding error: {e}")
            index, st_model = None, None

    # ✅ Cache safely
    cl.user_session.set("file_index", index)
    cl.user_session.set("file_text", "\n".join(texts) if texts else "")
    cl.user_session.set("st_model", st_model)
    cl.user_session.set("file_table", rows or [])
    cl.user_session.set("file_headers", headers or [])

    return index, texts, st_model


import json

# ------------------ Semantic Parser --------------------
async def parse_intent_semantic(user_question: str, headers: list, file_text: str):
    prompt = f"""
You are an assistant that interprets user questions about tabular data.

Available headers: {headers}
User question: "{user_question}"

Task:
1. Identify the intent(s): e.g. filter, count, list, average, sum, max, min, suggest, define.
2. Identify relevant headers.
3. Identify filter conditions (normalize values if needed).
Return JSON: {{ "actions":[], "headers":[], "filters":[] }}
    """
    response = await cl.make_async(run_qwen_sync)(prompt)
    try:
        return json.loads(response)
    except:
        return {"actions":["llm_fallback"],"headers":[],"filters":[]}

# ------------------ Universal Language Detection (cached) --------------------
async def detect_language(headers: list, file_text: str):
    cached_lang = cl.user_session.get("file_language")
    if cached_lang:
        return cached_lang

    prompt = f"""
Detect the primary language of this dataset.
Headers: {headers}
Sample text: {file_text[:500]}
Return ISO 639-1 code (e.g. en, fr, zh, ja, de, es, ru, hi).
    """
    response = await cl.make_async(run_qwen_sync)(prompt)
    lang_code = response.strip().lower()
    if len(lang_code) != 2:
        lang_code = "en"
    cl.user_session.set("file_language", lang_code)
    return lang_code

# ------------------ Localized Labels --------------------
def get_labels(language: str):
    labels = {
        "en":{"task":"Task Name","assigned":"Assigned To","start":"Start Date","end":"End Date","progress":"Progress"},
        "fr":{"task":"Tâche","assigned":"Attribué à","start":"Date début","end":"Date fin","progress":"Progression"},
        "zh":{"task":"任务","assigned":"负责人","start":"开始日期","end":"结束日期","progress":"进度"},
        "ja":{"task":"タスク","assigned":"担当者","start":"開始日","end":"終了日","progress":"進捗"},
        "de":{"task":"Aufgabe","assigned":"Zugewiesen an","start":"Startdatum","end":"Enddatum","progress":"Fortschritt"},
        "es":{"task":"Tarea","assigned":"Asignado a","start":"Fecha inicio","end":"Fecha fin","progress":"Progreso"},
        "ru":{"task":"Задача","assigned":"Назначено","start":"Дата начала","end":"Дата окончания","progress":"Прогресс"},
        "hi":{"task":"कार्य","assigned":"सौंपा गया","start":"प्रारंभ तिथि","end":"समाप्ति तिथि","progress":"प्रगति"}
    }
    return labels.get(language, labels["en"])

# ------------------ Universal Multi-Intent Handler --------------------
async def handle_user_query(msg_content: str, file_table: list, headers: list, file_text: str, file_path: str):
    # Direct keyword checks
    if "how many rows" in msg_content.lower():
        if file_table:
            return f"\n📊 Total rows: {len(file_table)}\n"
        else:
            row_count = len([line for line in file_text.splitlines() if line.strip()])
            return f"\n📊 Total rows (text-based): {row_count}\n"

    if "how many columns" in msg_content.lower():
        if headers:
            return f"\n📊 Total columns: {len(headers)}\n"
        else:
            first_line = next((line for line in file_text.splitlines() if line.strip()), "")
            col_count = len(set(first_line.split()))
            return f"\n📊 Estimated columns (text-based): {col_count}\n"

    # Semantic parsing
    parsed = await parse_intent_semantic(msg_content, headers, file_text)
    filters = parsed.get("filters", [])

    # Language detection + labels
    language = await detect_language(headers, file_text)
    labels = get_labels(language)
    combined_answer = ""

    # Apply structured filters
    structured_matches = []
    if file_table and filters:
        structured_matches = file_table
        for f in filters:
            for k, v in f.items():
                structured_matches = [
                    r for r in structured_matches
                    if str(r.get(k, "")).lower() == str(v).lower()
                ]

    if structured_matches:
        text = f"\n👷 Structured results ({filters}):\n"
        if headers:
            text += " | ".join(headers) + "\n" + "-"*70 + "\n"
        for r in structured_matches:
            text += " | ".join(str(r.get(h, "")) for h in headers) + "\n"
        combined_answer += text
        return combined_answer.strip()

    # Fall back to FAISS semantic search
    index = cl.user_session.get("file_index")
    st_model = cl.user_session.get("st_model")
    texts = cl.user_session.get("file_text", "").split("\n")

    if index and st_model and texts:
        query_vec = st_model.encode([msg_content])
        D, I = index.search(np.array(query_vec), k=3)
        if len(I[0]) > 0:
            context = "\n".join([texts[i] for i in I[0]])
            prompt = f"User question: {msg_content}\n\nRelevant context:\n{context}"
            llm_answer = await cl.make_async(run_qwen_sync)(prompt)
            combined_answer += f"\n{llm_answer}\n"
        else:
            # Fallback to chunking
            chunks = [file_text[i:i+4000] for i in range(0, len(file_text), 4000)]
            llm_answer = ""
            for chunk in chunks:
                prompt = f"User question: {msg_content}\n\nFile context:\n{chunk}"
                llm_answer += await cl.make_async(run_qwen_sync)(prompt) + "\n"
            combined_answer += f"\n{llm_answer}\n"

    return combined_answer.strip()


# ============================================================
# Utilities
# ============================================================
def run_qwen_sync(prompt: str) -> str:
    # Encode prompt into tokens
    input_ids = TOKENIZER.encode(prompt)

    # 🔑 Safe truncation by tokens
    max_context = MODEL.cfg["context_length"]
    if len(input_ids) > max_context:
        input_ids = input_ids[-max_context:]

    # Convert to tensor
    input_ids_tensor = torch.tensor(input_ids, device=DEVICE).unsqueeze(0)

    output = ""
    for tok in generate_text_basic_stream_cache(
        model=MODEL,
        token_ids=input_ids_tensor,
        max_new_tokens=MAX_NEW_TOKENS,
    ):
        token_id = tok.squeeze(0)
        if token_id in EOS_TOKEN_IDS:
            break
        output += TOKENIZER.decode(token_id.tolist())
    return output

class AgentState(TypedDict):
    messages: List[BaseMessage]

# ============================================================
# Chainlit Lifecycle
# ============================================================
@cl.password_auth_callback
def auth_callback(username: str, password: str):
    if (username, password) == ("admin", "admin"):
        return cl.User(identifier="admin", metadata={"role": "admin", "provider": "credentials"})
    return None

@cl.data_layer
def get_data_layer():
    conninfo = os.getenv("DATABASE_URL")
    if not conninfo:
        raise ValueError("DATABASE_URL not found in environment variables.")
    return SQLAlchemyDataLayer(conninfo=conninfo)

@cl.on_chat_resume
async def on_chat_resume(thread: ThreadDict):
    try:
        steps = thread.get("steps", [])
        messages = []
        for step in steps:
            step_type = step.get("type")
            content = (step.get("output") or "").strip()
            if not content:
                continue
            if step_type == "user_message":
                messages.append(HumanMessage(content=content))
            elif step_type == "assistant_message":
                messages.append(AIMessage(content=content))
        cl.user_session.set("state", {"messages": messages})
    except Exception as e:
        print(f"\nError resuming chat: {e}")
        cl.user_session.set("state", {"messages": []})

# ============================================================
# File Handling + Message
# ============================================================
@cl.on_message
async def on_message(msg: cl.Message):
    state = cl.user_session.get("state") or {"messages": []}
    cl.user_session.set("state", state)
    state["messages"].append(HumanMessage(content=msg.content))

    extracted_text = ""
    sidebar_elements = []
    file_path = None

    # --- Handle file attachments ---
    if msg.elements:
        for element in msg.elements:
            if element.type == "file":
                file_path = element.path
                file_name = element.name
                try:
                    # Use universal ingestion helper
                    index, texts, st_model = await prepare_file_context(file_path)
                    extracted_text = "\n".join(texts)
                    cl.user_session.set("file_text", extracted_text)

                    # Sidebar previews
                    if file_path.endswith(".pdf"):
                        sidebar_elements.append(cl.Pdf(path=file_path, name=file_name))
                    elif file_path.endswith(".docx"):
                        preview = "\n".join(texts)[:1000]
                        sidebar_elements.append(cl.Text(content=preview, name=file_name))
                    elif file_path.endswith(".xlsx") or file_path.endswith(".csv"):
                        preview = "\n".join(texts)[:1000]
                        sidebar_elements.append(cl.Text(content=preview, name=file_name))
                    elif file_path.lower().endswith((".png", ".jpg", ".jpeg")):
                        sidebar_elements.append(cl.Image(path=file_path, name=file_name))

                    if extracted_text:
                        preview = extracted_text[:1000]
                        state["messages"].append(HumanMessage(content=f"[File {file_name} content]: {preview}..."))

                except Exception as e:
                    print(f"⚠️ Error reading file {file_name}: {e}")

    # --- Show files in sidebar ---
    if sidebar_elements:
        await cl.ElementSidebar.set_elements(sidebar_elements)
        await cl.ElementSidebar.set_title("Uploaded Files")

    # --- Retrieve stored context ---
    file_table = cl.user_session.get("file_table")
    headers = cl.user_session.get("file_headers", [])
    file_text = cl.user_session.get("file_text", "")

    # --- Answer user query ---
    if file_table:
        detected_col = auto_detect_column(headers, msg.content)
        if detected_col:
            keywords = extract_keywords(msg.content, headers, file_table)
            answer = get_filtered_table(file_table, detected_col, keywords)
            if not answer or "No results" in answer:
                answer = await handle_user_query(msg.content, file_table, headers, file_text, file_path)
        else:
            answer = await handle_user_query(msg.content, file_table, headers, file_text, file_path)
    elif file_text:
        # Use semantic handler if text exists but no structured table
        answer = await handle_user_query(msg.content, [], [], file_text, file_path)
    else:
        # Pure LLM response
        answer = await cl.make_async(run_qwen_sync)(msg.content)

    await cl.Message(content=answer).send()


@cl.on_stop
def on_stop():
    print("The user wants to stop the task!")

@cl.on_chat_end
def on_chat_end():
    print("The user disconnected!")

# ------------------- Starters -------------------
@cl.set_starters
async def set_starters():
    return [
        cl.Starter(
            label="Morning routine ideation",
            message="Can you help me create a personalized morning routine that would help increase my productivity throughout the day? Start by asking me about my current habits and what activities energize me in the morning.",
            icon="/public/idea.png"
        ),
        cl.Starter(
            label="Explain superconductors",
            message="Explain superconductors like I'm five years old.",
            icon="/public/learn.png"
        ),
        cl.Starter(
            label="Python script for daily email reports",
            message="Write a script to automate sending daily email reports in Python, and walk me through how I would set it up.",
            icon="/public/terminal.png",
            command="code"
        ),
        cl.Starter(
            label="Text inviting friend to wedding",
            message="Write a text asking a friend to be my plus-one at a wedding next month. I want to keep it super short and casual, and offer an out.",
            icon="/public/write.png"
        )
    ]

def register_tools(app):
    if os.getenv("STRIPE_ENABLED", "false").lower() == "true":
        from stripe_mcp import StripeMCP
        app.register_tool(StripeMCP())
        print("✅ Stripe MCP enabled.")
    else:
        print("🚫 Stripe MCP disabled — prompt-only mode.")

# Register tools based on environment flag
register_tools(cl)
