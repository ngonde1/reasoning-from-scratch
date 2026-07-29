import os
import json
import re
import datetime
import openpyxl
import csv
import docx
import PyPDF2
from pdf2image import convert_from_path
import pytesseract
from collections import defaultdict

# ------------------- Utility -------------------

def sanitize_row(row: dict) -> dict:
    """Convert non-serializable values (like datetime) into strings."""
    clean_row = {}
    for k, v in row.items():
        if isinstance(v, (datetime.date, datetime.datetime)):
            clean_row[k] = v.isoformat()
        elif v is None:
            clean_row[k] = ""
        else:
            clean_row[k] = v
    return clean_row

# ------------------- File Loaders -------------------

def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [str(cell).strip() if cell else f"Column{i}" for i, cell in enumerate(next(sheet.iter_rows(values_only=True)))]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        rows.append(sanitize_row(row_dict))
    return headers, rows

def load_csv(file_path):
    rows, headers = [], []
    with open(file_path, newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        for row in reader:
            rows.append(sanitize_row(row))
    return headers, rows

def load_docx(file_path):
    doc = docx.Document(file_path)
    texts = [para.text for para in doc.paragraphs if para.text.strip()]
    return texts

def load_pdf(file_path):
    texts = []
    reader = PyPDF2.PdfReader(file_path)
    POPPLER_PATH = r"C:\poppler-26.02.0\Library\bin"
    for page_num, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text()
            if page_text and page_text.strip():
                texts.append(page_text)
            else:
                images = convert_from_path(file_path, first_page=page_num+1, last_page=page_num+1, poppler_path=POPPLER_PATH)
                for img in images:
                    ocr_text = pytesseract.image_to_string(img)
                    if ocr_text.strip():
                        texts.append(ocr_text)
        except Exception as e:
            texts.append(f"⚠️ OCR error on page {page_num+1}: {e}")
    return texts if texts else ["⚠️ No text extracted from PDF."]

def load_txt(file_path):
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        file_text = f.read()
    return file_text.splitlines()

def load_json(file_path):
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        data = json.load(f)
    if isinstance(data, list) and all(isinstance(item, dict) for item in data):
        headers = list(data[0].keys()) if data else []
        rows = [sanitize_row(r) for r in data]
        return headers, rows
    else:
        file_text = json.dumps(data, indent=2, default=str)
        return [], [{"RawJSON": file_text}]

# ------------------- Analysis Helpers -------------------

def auto_detect_columns(headers, rows):
    categorical_cols, numeric_cols, date_cols = [], [], []
    for h in headers:
        values = [str(r.get(h, "")).strip() for r in rows if r.get(h)]
        if not values:
            continue
        numeric_count, date_count = 0, 0
        for v in values[:20]:
            try:
                float(v); numeric_count += 1
            except:
                try:
                    datetime.datetime.fromisoformat(v); date_count += 1
                except:
                    pass
        if numeric_count > len(values) * 0.5:
            numeric_cols.append(h)
        elif date_count > len(values) * 0.3:
            date_cols.append(h)
        else:
            categorical_cols.append(h)
    return categorical_cols, numeric_cols, date_cols

def build_copy_table(rows, headers, limit=30):
    if not rows:
        return "⚠️ No matching records found."
    table = "| " + " | ".join(headers) + " |\n"
    table += "| " + " | ".join(["---"] * len(headers)) + " |\n"
    for r in rows[:limit]:
        row_values = [str(r.get(h, "")) for h in headers]
        table += "| " + " | ".join(row_values) + " |\n"
    return table

def compute_all_numeric_stats(rows, headers):
    numeric_stats = {}
    for col in headers:
        values = []
        for r in rows:
            val = r.get(col)
            try:
                num = float(val); values.append(num)
            except (TypeError, ValueError):
                continue
        if values:
            numeric_stats[col] = {
                "count": len(values),
                "sum": sum(values),
                "avg": sum(values)/len(values),
                "min": min(values),
                "max": max(values),
            }
    if not numeric_stats:
        return "⚠️ No numeric columns found."
    result = "📊 **Numeric Stats:**\n\n"
    for col, stats in numeric_stats.items():
        result += f"- {col}: Count={stats['count']}, Sum={stats['sum']:.2f}, Avg={stats['avg']:.2f}, Min={stats['min']:.2f}, Max={stats['max']:.2f}\n"
    return result

# ------------------- Universal Response -------------------

async def professional_llm_response(
    user_query: str,
    file_table: list,
    headers: list,
    llm_runner,
    file_text: str = "",
    preview_limit: int = 30
) -> str:
    if not file_table and not file_text:
        return await llm_runner(f"User question: {user_query}\n\n⚠️ No file content available.")

    filtered_rows = file_table or []
    total_matches = len(filtered_rows)

    json_summary = [{h: row.get(h, "") for h in headers} for row in filtered_rows]

    categorical_cols, numeric_cols, date_cols = auto_detect_columns(headers, filtered_rows)

    numeric_stats = compute_all_numeric_stats(filtered_rows, headers) if numeric_cols else ""

    group_by_summary = ""
    if categorical_cols and numeric_cols:
        group_by_summary = "📊 **Group-by Aggregation:**\n\n"
        for cat_col in categorical_cols:
            for num_col in numeric_cols:
                groups = defaultdict(list)
                for row in filtered_rows:
                    cat_val = row.get(cat_col) or "Unknown"
                    try:
                        num_val = float(row.get(num_col, 0))
                        groups[cat_val].append(num_val)
                    except:
                        continue
                if groups:
                    group_by_summary += f"- {num_col} by {cat_col}:\n\n"
                    group_by_summary += "| Group | Average |\n| --- | --- |\n"
                    for g, vals in groups.items():
                        avg = sum(vals)/len(vals)
                        group_by_summary += f"| {g} | {avg:.2f} |\n"
                    group_by_summary += "\n"

    trend_summary = ""
    if numeric_cols and date_cols and categorical_cols:
        trend_summary = "📈 **Trend Detection:**\n\n"
        for cat_col in categorical_cols:
            for num_col in numeric_cols:
                for date_col in date_cols:
                    dept_trends = defaultdict(list)
                    for row in filtered_rows:
                        cat_val = row.get(cat_col) or "Unknown"
                        try:
                            num_val = float(row.get(num_col, 0))
                            date_val = row.get(date_col)
                            if isinstance(date_val, str):
                                try:
                                    date_val = datetime.datetime.fromisoformat(date_val)
                                except:
                                    continue
                            if isinstance(date_val, datetime.date):
                                dept_trends[cat_val].append((date_val, num_val))
                        except:
                            continue
                    for dept, values in dept_trends.items():
                        values.sort(key=lambda x: x[0])
                        if len(values) >= 2:
                            first, last = values[0][1], values[-1][1]
                            if last > first:
                                trend_summary += f"- {dept} ({num_col}): Improving ({first:.2f} → {last:.2f})\n"
                            elif last < first:
                                trend_summary += f"- {dept} ({num_col}): Declining ({first:.2f} → {last:.2f})\n"
                            else:
                                trend_summary += f"- {dept} ({num_col}): Stable ({first:.2f})\n"

    enrichment = ""
    text_blob = " ".join(headers).lower() + " " + user_query.lower() + " " + (file_text.lower() if file_text else "")
