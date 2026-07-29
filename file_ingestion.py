
# Not important for now


import os
import csv
import re
import json
import openpyxl
import docx
import PyPDF2
from pathlib import Path
from pdf2image import convert_from_path
from PIL import Image as PILImage
import pytesseract
from difflib import get_close_matches
from sqlalchemy import text as sql_text
import chainlit as cl

# ------------------- Persistence -------------------

from reasoning_from_scratch.rag_utils import add_to_index
from reasoning_from_scratch.utils import sanitize_row, chunk_rows, chunk_text, ingest_file, get_data_layer

async def persist_file_content(file_name_or_path, headers=None, rows=None, text_content=None):
    """
    Unified persistence function:
    - Accepts either a file path (new style) OR legacy args (file_name, headers, rows, text_content).
    - Handles chunking automatically for structured files.
    - Indexes both structured rows and unstructured text into FAISS.
    - Saves normalized payloads into Postgres.
    """
    data_layer = get_data_layer()
    thread_id = cl.context.session.thread_id

    # ✅ If only a file path is passed (new style)
    if headers is None and rows is None and text_content is None:
        file_path = Path(file_name_or_path)
        file_name = file_path.name
        headers, rows_or_text = ingest_file(str(file_path))

        if headers and rows_or_text:  # structured file
            for chunk in chunk_rows(rows_or_text, chunk_size=int(os.getenv("ROW_CHUNK_SIZE", 500))):
                payload = {
                    "thread_id": thread_id,
                    "file_name": file_name,
                    "headers": json.dumps(headers),
                    "rows": json.dumps([sanitize_row(r) for r in chunk]),
                    "text_content": None,
                }
                async with data_layer.engine.begin() as conn:
                    await conn.execute(
                        sql_text("""INSERT INTO file_contents (thread_id, file_name, headers, rows, text_content)
                                    VALUES (:thread_id, :file_name, :headers, :rows, :text_content)"""),
                        payload
                    )
                add_to_index(chunk, headers)

        else:  # unstructured file
            text_content = rows_or_text
            payload = {
                "thread_id": thread_id,
                "file_name": file_name,
                "headers": None,
                "rows": None,
                "text_content": text_content if isinstance(text_content, str) else json.dumps(text_content),
            }
            async with data_layer.engine.begin() as conn:
                await conn.execute(
                    sql_text("""INSERT INTO file_contents (thread_id, file_name, headers, rows, text_content)
                                VALUES (:thread_id, :file_name, :headers, :rows, :text_content)"""),
                    payload
                )
            # Index text chunks
            if isinstance(text_content, str):
                for chunk in chunk_text(text_content, chunk_size=int(os.getenv("TEXT_CHUNK_SIZE", 1000)),
                                        overlap=int(os.getenv("TEXT_CHUNK_OVERLAP", 100))):
                    add_to_index([{"RawText": chunk}], ["RawText"])
            elif isinstance(text_content, list):
                for chunk in text_content:
                    for sub_chunk in chunk_text(chunk, chunk_size=int(os.getenv("TEXT_CHUNK_SIZE", 1000)),
                                                overlap=int(os.getenv("TEXT_CHUNK_OVERLAP", 100))):
                        add_to_index([{"RawText": sub_chunk}], ["RawText"])

    # ✅ If legacy args are passed (old style)
    else:
        file_name = file_name_or_path
        payload = {
            "thread_id": thread_id,
            "file_name": file_name,
            "headers": json.dumps(headers) if headers else None,
            "rows": json.dumps([sanitize_row(r) for r in rows]) if rows else None,
            "text_content": text_content,
        }
        async with data_layer.engine.begin() as conn:
            await conn.execute(
                sql_text("""INSERT INTO file_contents (thread_id, file_name, headers, rows, text_content)
                            VALUES (:thread_id, :file_name, :headers, :rows, :text_content)"""),
                payload
            )
        if rows and headers:
            add_to_index(rows, headers)
        if text_content:
            for chunk in chunk_text(text_content, chunk_size=int(os.getenv("TEXT_CHUNK_SIZE", 1000)),
                                    overlap=int(os.getenv("TEXT_CHUNK_OVERLAP", 100))):
                add_to_index([{"RawText": chunk}], ["RawText"])

    print(f"✅ Saved {file_name_or_path} to Postgres and FAISS for thread {thread_id}")

# ------------------- Loaders -------------------
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = []
    for i, cell in enumerate(next(sheet.iter_rows(values_only=True))):
        headers.append(str(cell).strip() if cell else f"Column{i}")
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        rows.append(row_dict)
    return headers, rows

def load_csv(file_path):
    rows, headers = [], []
    with open(file_path, newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        for row in reader:
            rows.append(row)
    return headers, rows

def load_docx(file_path):
    doc = docx.Document(file_path)
    texts = [para.text for para in doc.paragraphs if para.text.strip()]
    return texts

def load_pdf(file_path):
    texts = []
    reader = PyPDF2.PdfReader(file_path)
    POPPLER_PATH = r"C:\poppler-26.02.0\Library\bin"

    for page_num, page in enumerate(reader.pages):
        try:
            page_text = page.extract_text()
            if page_text and page_text.strip():
                texts.append(page_text)
            else:
                images = convert_from_path(
                    file_path,
                    first_page=page_num + 1,
                    last_page=page_num + 1,
                    poppler_path=POPPLER_PATH
                )
                for img in images:
                    ocr_text = pytesseract.image_to_string(img)
                    if ocr_text.strip():
                        texts.append(ocr_text)
        except Exception as e:
            print(f"⚠️ OCR failed on page {page_num+1}: {e}")
            texts.append(f"⚠️ OCR error on page {page_num+1}")

    return texts if texts else ["⚠️ No text extracted from PDF."]

def load_txt(file_path):
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        file_text = f.read()
    return file_text.splitlines()

def load_json(file_path):
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        data = json.load(f)
    if isinstance(data, list) and all(isinstance(item, dict) for item in data):
        headers = list(data[0].keys()) if data else []
        rows = data
        return headers, rows
    else:
        file_text = json.dumps(data, indent=2)
        return [], [{"RawJSON": file_text}]

# ------------------- Helpers -------------------
def build_table(rows, headers, limit=20):
    if not rows:
        return "⚠️ No matching records found."
    table = "| " + " | ".join(headers) + " |\n"
    table += "| " + " | ".join(["---"] * len(headers)) + " |\n"
    for r in rows[:limit]:
        row_values = [str(r.get(h, "")) for h in headers]
        table += "| " + " | ".join(row_values) + " |\n"
    return table

def parse_numeric(value):
    if value is None:
        return None
    raw_text = str(value).strip()
    raw_text = re.sub(r"(USD|EUR|GBP|JPY|CFA|NGN|INR|CAD|AUD|CHF|₦|₿|¥|€|£|\$)", "", raw_text, flags=re.IGNORECASE)
    raw_text = re.sub(r"[,%]", "", raw_text)
    raw_text = re.sub(r"[A-Za-z:]", "", raw_text)
    cleaned = re.sub(r"[^\d\.\-]", "", raw_text)
    try:
        return float(cleaned)
    except ValueError:
        return None

def filter_rows_by_numeric(rows, column_name, threshold, operator=">"):
    results = []
    for r in rows:
        val = parse_numeric(r.get(column_name))
        if val is None:
            continue
        if operator == ">" and val > threshold:
            results.append(r)
        elif operator == "<" and val < threshold:
            results.append(r)
        elif operator == ">=" and val >= threshold:
            results.append(r)
        elif operator == "<=" and val <= threshold:
            results.append(r)
        elif operator == "==" and val == threshold:
            results.append(r)
    return results

def compute_all_numeric_stats(rows, headers):
    numeric_stats = {}
    for col in headers:
        values = []
        for r in rows:
            try:
                num = float(r.get(col))
                values.append(num)
            except (TypeError, ValueError):
                continue
        if values:
            numeric_stats[col] = {
                "count": len(values),
                "sum": sum(values),
                "avg": sum(values) / len(values),
                "min": min(values),
                "max": max(values),
            }
    if not numeric_stats:
        return "⚠️ No numeric columns found."
    result = "📊 **Numeric Stats:**\n\n"
    for col, stats in numeric_stats.items():
        result += (
            f"- {col}: Count={stats['count']}, Sum={stats['sum']:.2f}, "
            f"Avg={stats['avg']:.2f}, Min={stats['min']:.2f}, Max={stats['max']:.2f}\n"
        )
    return result
