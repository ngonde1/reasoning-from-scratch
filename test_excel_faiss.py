import openpyxl
import faiss
import numpy as np
from sentence_transformers import SentenceTransformer
from sentence_transformers.sentence_transformer import modules

def detect_headers(sheet, max_scan=10):
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
        if len(non_empty) >= 2:  # likely header row
            headers = [str(cell).strip() if cell else f"Column{j}" for j, cell in enumerate(row)]
            return i, headers
    return 1, []

def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    header_row_index, headers = detect_headers(sheet)

    # Drop Column0 if it's just a placeholder
    if headers and headers[0].startswith("Column"):
        headers = headers[1:]

    rows = []
    for row in sheet.iter_rows(min_row=header_row_index+1, values_only=True):
        if not any(row):
            continue
        row = row[1:]  # drop first empty cell
        row_dict = {}
        for i, header in enumerate(headers):
            value = row[i] if i < len(row) else None
            row_dict[header] = value
        rows.append(row_dict)

    return headers, rows

def build_faiss_index(rows):
    texts = [" | ".join(f"{k}: {v}" for k, v in r.items() if v is not None) for r in rows]

    # ✅ Manually construct model with updated API
    word_embedding_model = modules.Transformer("all-MiniLM-L6-v2")
    pooling_model = modules.Pooling(word_embedding_model.get_embedding_dimension())
    model = SentenceTransformer(modules=[word_embedding_model, pooling_model])

    embeddings = model.encode(texts)

    dimension = embeddings.shape[1]
    index = faiss.IndexFlatL2(dimension)
    index.add(np.array(embeddings))

    return index, model, texts

def query_index(index, model, texts, rows, query, k=5, filters=None):
    query_vec = model.encode([query])
    D, I = index.search(np.array(query_vec), k)

    print(f"\nQuery: {query}")
    print("Top matches:")

    # Pass 1: structured filter across ALL rows
    structured_matches = []
    if filters:
        for r in rows:
            if all(f(r) for f in filters):
                structured_matches.append(r)

    if structured_matches:
        print("(Structured matches found)")
        for r in structured_matches[:k]:
            print("-", r)
    else:
        print("(No structured matches, showing semantic results instead)")
        for idx in I[0]:
            print("-", rows[idx])

def main():
    file_path = r"C:\Users\Immanuel\Desktop\Project-Management-Sample-Data.xlsx"
    headers, rows = load_excel(file_path)
    print("Headers:", headers)
    print("Total rows:", len(rows))
    print("First row:", rows[0])

    index, model, texts = build_faiss_index(rows)

    # Pure semantic query
    query_index(index, model, texts, rows, "Who is in Engineering?")

    # Hybrid semantic + structured filter (single condition)
    query_index(index, model, texts, rows, "Show tasks with 100% progress",
                filters=[lambda r: r.get("Progress") == 1.0])

    # Compound structured filter (multiple conditions)
    query_index(index, model, texts, rows,
                "Engineering tasks with 100% progress",
                filters=[
                    lambda r: r.get("Progress") == 1.0,
                    lambda r: r.get("Project Name") == "Engineering"
                ])

if __name__ == "__main__":
    main()
