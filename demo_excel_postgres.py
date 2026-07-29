import json
import psycopg2
import openpyxl
from datetime import datetime

# 1. Load Excel rows (spaghetti)
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    headers = [str(cell).strip() for cell in next(sheet.iter_rows(values_only=True))]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i in range(len(headers)):
            val = row[i]
            if isinstance(val, datetime):
                val = val.strftime("%Y-%m-%d")
            row_dict[headers[i]] = val
        rows.append(row_dict)
    return headers, rows

# 2. Save into Postgres
def save_to_postgres(headers, rows):
    conn = psycopg2.connect(
        dbname="my_chainlit_db",
        user="postgres",
        password="kevin",  # change to your real password
        host="localhost",
        port="5432"
    )
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO file_contents (thread_id, file_name, headers, rows, text_content)
        VALUES (%s, %s, %s, %s, %s)
        """,
        (
            "demo-thread-123",
            "Employee Sample Data.xlsx",
            json.dumps(headers),
            json.dumps(rows),
            None
        )
    )
    conn.commit()
    cur.close()
    conn.close()
    print("✅ Saved Excel rows as JSON smoothies")

# 3. Load back from Postgres
def load_from_postgres():
    conn = psycopg2.connect(
        dbname="my_chainlit_db",
        user="postgres",
        password="kevin",
        host="localhost",
        port="5432"
    )
    cur = conn.cursor()
    cur.execute("SELECT headers, rows FROM file_contents WHERE file_name = %s", ("Employee Sample Data.xlsx",))
    headers_json, rows_json = cur.fetchone()
    headers = json.loads(headers_json) if isinstance(headers_json, str) else headers_json
    rows = json.loads(rows_json) if isinstance(rows_json, str) else rows_json
    cur.close()
    conn.close()
    print("✅ Loaded back into Python objects")
    return headers, rows

# 4. Filter employees with salary > threshold
def filter_high_salary(rows, threshold=70000):
    return [r for r in rows if r.get("Annual Salary", 0) and r["Annual Salary"] > threshold]

# 5. Print clean summary table
def print_summary_table(employees):
    if not employees:
        print("⚠️ No employees found above threshold.")
        return
    # Build Markdown-style table
    print("\n📊 Summary Table (Name, Department, Salary, Bonus %):\n")
    print("| Full Name | Department | Annual Salary | Bonus % |")
    print("|-----------|------------|---------------|---------|")
    for emp in employees:
        name = emp.get("Full Name", "")
        dept = emp.get("Department", "")
        salary = emp.get("Annual Salary", "")
        bonus = emp.get("Bonus %", "")
        print(f"| {name} | {dept} | {salary} | {bonus} |")

# Run demo
headers, rows = load_excel(r"C:\Users\Immanuel\Desktop\reasoning-from-scratch\Employee Sample Data.xlsx")
save_to_postgres(headers, rows)
headers_back, rows_back = load_from_postgres()

high_salary_employees = filter_high_salary(rows_back, 70000)
print_summary_table(high_salary_employees)
